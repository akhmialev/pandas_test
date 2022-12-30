from django.core.management.base import BaseCommand
from prj.settings import DATA_DIR
from openpyxl import load_workbook
from market.models import Category, Product

class Command(BaseCommand):

    def handle(self, *args, **options):
        print('Clearing DB')
        Category.objects.all().delete()
        Product.objects.all().delete()

        print('Start importing from excel %s' % DATA_DIR)

        wb = load_workbook(DATA_DIR+'/data.xlsx')
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

        cat = None
        for cnt in range(1, sheet.max_row+1):
            item = sheet.cell(row=cnt, column=1).value
            none_item = sheet.cell(row=cnt, column=2).value

            if none_item == None:
                print('Create category')
                cat = Category()
                cat.name = item
                cat.save()
            else:
                print('Create product')
                if cat:
                    p = Product()
                    p.name = item
                    p.category = cat
                    p.save()
